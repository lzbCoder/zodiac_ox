import asyncio, csv, io
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.ext.asyncio import AsyncSession
from database import get_db
from schemas.rag_eval import (
    RagEvalDatasetCreate, RagEvalDatasetUpdate, RagEvalDatasetResponse,
    RagEvalQuestionResponse,
    RagEvalTaskCreate, RagEvalTaskResponse, RagEvalTaskUpdate,
    RagEvalResultResponse, RagEvalReportData,
    RagEvalConfigUpdate, RagEvalConfigResponse,
    RagEvalLabelTaskCreate, RagEvalLabelTaskUpdate, RagEvalLabelTaskResponse,
    RagEvalLabelDetailResponse, RagEvalLabelDetailSave, RagEvalLabelBatchSave,
    ChunkCandidatesResponse, LabelExportResponse,
)
from services import eval_dataset_service, eval_import_service, eval_config_service, eval_task_service, eval_engine, eval_label_service

router = APIRouter(prefix="/api/rag/eval", tags=["RAG评测中心"])


# ── Dataset CRUD ──────────────────────────────────────

@router.get("/datasets")
async def list_datasets(
    kb_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_dataset_service.list_datasets(db, kb_id, page, page_size)
    return {"items": [RagEvalDatasetResponse.model_validate(d) for d in items], "total": total, "page": page, "page_size": page_size}


@router.post("/datasets", response_model=RagEvalDatasetResponse)
async def create_dataset(data: RagEvalDatasetCreate, db: AsyncSession = Depends(get_db)):
    return await eval_dataset_service.create_dataset(db, data.kb_id, data.name, data.description, data.created_by)


@router.get("/datasets/{dataset_id}", response_model=RagEvalDatasetResponse)
async def get_dataset(dataset_id: int, db: AsyncSession = Depends(get_db)):
    ds = await eval_dataset_service.get_dataset(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="评测集不存在")
    return ds


@router.put("/datasets/{dataset_id}", response_model=RagEvalDatasetResponse)
async def update_dataset(dataset_id: int, data: RagEvalDatasetUpdate, db: AsyncSession = Depends(get_db)):
    ds = await eval_dataset_service.update_dataset(db, dataset_id, data.name, data.description)
    if not ds:
        raise HTTPException(status_code=404, detail="评测集不存在")
    return ds


@router.delete("/datasets/{dataset_id}")
async def delete_dataset(dataset_id: int, db: AsyncSession = Depends(get_db)):
    ok = await eval_dataset_service.delete_dataset(db, dataset_id)
    if not ok:
        raise HTTPException(status_code=404, detail="评测集不存在")
    return {"message": "已删除"}


# ── Questions ──────────────────────────────────────────

@router.get("/datasets/{dataset_id}/questions")
async def list_questions(
    dataset_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_dataset_service.list_questions(db, dataset_id, page, page_size)
    return {"items": [RagEvalQuestionResponse.model_validate(q) for q in items], "total": total, "page": page, "page_size": page_size}


@router.post("/datasets/{dataset_id}/import")
async def import_questions(
    dataset_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    ds = await eval_dataset_service.get_dataset(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="评测集不存在")

    try:
        rows = await eval_import_service.parse_upload_file(file)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ImportError as e:
        raise HTTPException(status_code=500, detail=str(e))

    if not rows:
        raise HTTPException(status_code=400, detail="文件中无有效数据")

    count = await eval_dataset_service.import_questions_from_rows(db, dataset_id, ds.kb_id, rows)
    return {"message": f"成功导入 {count} 条问题", "count": count}


@router.delete("/questions/{question_id}")
async def delete_question(question_id: int, db: AsyncSession = Depends(get_db)):
    ok = await eval_dataset_service.delete_question(db, question_id)
    if not ok:
        raise HTTPException(status_code=404, detail="问题不存在")
    return {"message": "已删除"}


# ── Tasks ──────────────────────────────────────────────

@router.get("/tasks")
async def list_tasks(
    kb_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_task_service.list_tasks(db, kb_id, page, page_size)
    return {"items": [RagEvalTaskResponse.model_validate(t) for t in items], "total": total, "page": page, "page_size": page_size}


@router.post("/tasks", response_model=RagEvalTaskResponse)
async def create_task(data: RagEvalTaskCreate, db: AsyncSession = Depends(get_db)):
    # Validate: manual tasks need a dataset
    if data.task_type != "chat_sample":
        if not data.dataset_id:
            raise HTTPException(status_code=400, detail="传统标注评测需要选择评测集")
        ds = await eval_dataset_service.get_dataset(db, data.dataset_id)
        if not ds:
            raise HTTPException(status_code=404, detail="评测集不存在")

    # Fill defaults from KB config
    config = await eval_config_service.get_or_create_config(db, data.kb_id)
    top_k = data.top_k if data.top_k != 5 else config.default_top_k
    retriever_mode = data.retriever_mode if data.retriever_mode != "normal" else config.default_retriever_mode

    # Auto-fill RAGAS model defaults from system config if not specified
    from cache.model_config_cache import get_ragas_default_answer_model, get_ragas_default_eval_model
    model_name = data.model_name or get_ragas_default_answer_model()
    eval_model = data.eval_model or get_ragas_default_eval_model()

    return await eval_task_service.create_task(
        db, data.name, data.dataset_id, data.kb_id, top_k, retriever_mode,
        model_name, data.enable_ragas, eval_model,
        task_type=data.task_type,
        sample_time_start=data.sample_time_start,
        sample_time_end=data.sample_time_end,
        sample_count=data.sample_count,
        sample_strategy=data.sample_strategy,
    )


@router.post("/tasks/{task_id}/run")
async def run_task(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await eval_task_service.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    if task.status not in ("pending", "failed"):
        raise HTTPException(status_code=400, detail=f"任务状态为 {task.status}，无法启动")

    # Launch background evaluation
    asyncio.create_task(eval_engine.run_evaluation(task.id))
    return {"message": "评测任务已启动", "task_id": task_id}


@router.post("/tasks/{task_id}/cancel")
async def cancel_task(task_id: int, db: AsyncSession = Depends(get_db)):
    ok = await eval_task_service.cancel_task(db, task_id)
    if not ok:
        raise HTTPException(status_code=400, detail="任务状态不允许取消")
    eval_engine.request_cancel(task_id)
    return {"message": "已请求取消"}


@router.put("/tasks/{task_id}", response_model=RagEvalTaskResponse)
async def update_task(task_id: int, data: RagEvalTaskUpdate, db: AsyncSession = Depends(get_db)):
    task = await eval_task_service.update_task(db, task_id, data.name)
    if not task:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    return task


@router.delete("/tasks/{task_id}")
async def delete_task(task_id: int, db: AsyncSession = Depends(get_db)):
    ok = await eval_task_service.delete_task(db, task_id)
    if not ok:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    return {"message": "已删除"}


@router.get("/tasks/{task_id}")
async def get_task(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await eval_task_service.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    return RagEvalTaskResponse.model_validate(task)


# ── Reports ────────────────────────────────────────────

@router.get("/tasks/{task_id}/report")
async def get_report(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await eval_task_service.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    results = await eval_task_service.get_results(db, task_id)
    return RagEvalReportData(
        task=RagEvalTaskResponse.model_validate(task),
        total_questions=len(results),
        results=[RagEvalResultResponse.model_validate(r) for r in results],
    )


@router.get("/tasks/{task_id}/results")
async def get_results(
    task_id: int,
    hit_filter: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    results = await eval_task_service.get_results(db, task_id, hit_filter)
    return {"items": [RagEvalResultResponse.model_validate(r) for r in results], "total": len(results)}


@router.get("/tasks/{task_id}/report/export")
async def export_report(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await eval_task_service.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="评测任务不存在")
    results = await eval_task_service.get_results(db, task_id)

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评测报告"
    if task.enable_ragas:
        ws.append(["问题ID", "问题", "召回率", "准确率", "是否命中", "排名", "MRR", "检索耗时(s)",
                    "模型回答", "生成耗时(s)", "上下文精确度", "上下文召回率", "忠实度", "答案相关度"])
        for r in results:
            ws.append([r.qid, r.query, r.recall, r.precision, r.hit, r.rank, r.mrr, r.retrieve_time,
                       r.answer, r.answer_time, r.context_precision, r.context_recall,
                       r.faithfulness, r.answer_relevancy])
    else:
        ws.append(["问题ID", "问题", "召回率", "准确率", "是否命中", "排名", "MRR", "检索耗时(s)"])
        for r in results:
            ws.append([r.qid, r.query, r.recall, r.precision, r.hit, r.rank, r.mrr, r.retrieve_time])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=eval_report_{task_id}.xlsx"},
    )


@router.get("/reports")
async def list_completed_reports(
    kb_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_task_service.get_completed_tasks(db, kb_id, page, page_size)
    return {"items": [RagEvalTaskResponse.model_validate(t) for t in items], "total": total, "page": page, "page_size": page_size}


# ── Config ─────────────────────────────────────────────

@router.get("/configs/{kb_id}", response_model=RagEvalConfigResponse)
async def get_config(kb_id: int, db: AsyncSession = Depends(get_db)):
    return await eval_config_service.get_or_create_config(db, kb_id)


@router.put("/configs/{kb_id}", response_model=RagEvalConfigResponse)
async def update_config(kb_id: int, data: RagEvalConfigUpdate, db: AsyncSession = Depends(get_db)):
    return await eval_config_service.update_config(db, kb_id, data.default_top_k, data.default_retriever_mode)


# ══════════════════════════════════════════════════════════
#  Label Tool
# ══════════════════════════════════════════════════════════

# ── File Parse ────────────────────────────────────────────

@router.post("/label/parse-file")
async def parse_label_file(file: UploadFile = File(...)):
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ("csv", "xlsx", "xls", "json"):
        raise HTTPException(status_code=400, detail=f"不支持的文件格式：{ext}，仅支持 Excel(.xlsx/.xls)、JSON(.json)、CSV(.csv)")

    content = await file.read()

    try:
        if ext == "json":
            import json as _json
            data = _json.loads(content.decode("utf-8-sig"))
            if not isinstance(data, list):
                raise ValueError("JSON文件应为对象数组")
            items = []
            for i, item in enumerate(data):
                q = str(item.get("query", "")).strip()
                if not q:
                    continue
                items.append({
                    "query": q,
                    "standard_answer": str(item.get("standard_answer", "")).strip() or None,
                })
            return {"items": items, "count": len(items)}

        elif ext in ("xlsx", "xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl未安装")
            wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            if not ws:
                raise HTTPException(status_code=400, detail="Excel文件为空")
            headers = [str(h).strip() if h else "" for h in next(ws.iter_rows(values_only=True), [])]
            if "query" not in headers:
                raise HTTPException(status_code=400, detail=f"Excel缺少query列，当前表头：{headers}")
            items = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(headers, row))
                q = str(d.get("query", "")).strip()
                if not q:
                    continue
                items.append({
                    "query": q,
                    "standard_answer": str(d.get("standard_answer", "")).strip() or None,
                })
            wb.close()
            return {"items": items, "count": len(items)}

        else:  # csv
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                raise HTTPException(status_code=400, detail="CSV文件为空")
            headers = [h.strip().lower() for h in rows[0]]
            if "query" not in headers:
                raise HTTPException(status_code=400, detail=f"CSV缺少query列，当前表头：{headers}")
            q_idx = headers.index("query")
            a_idx = headers.index("standard_answer") if "standard_answer" in headers else -1
            items = []
            for row in rows[1:]:
                if len(row) <= q_idx:
                    continue
                q = row[q_idx].strip()
                if not q:
                    continue
                items.append({
                    "query": q,
                    "standard_answer": row[a_idx].strip() if a_idx >= 0 and len(row) > a_idx else None,
                })
            return {"items": items, "count": len(items)}

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败：{str(e)}")


# ── Label Tasks ───────────────────────────────────────────

@router.get("/label/tasks")
async def list_label_tasks(
    kb_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_label_service.list_label_tasks(db, kb_id, page, page_size)
    return {"items": [RagEvalLabelTaskResponse.model_validate(t) for t in items], "total": total, "page": page, "page_size": page_size}


@router.post("/label/tasks", response_model=RagEvalLabelTaskResponse)
async def create_label_task(data: RagEvalLabelTaskCreate, db: AsyncSession = Depends(get_db)):
    # Check for duplicate queries within the import
    seen = set()
    unique = []
    for q in data.queries:
        if q.query.strip() in seen:
            continue
        seen.add(q.query.strip())
        unique.append({"query": q.query.strip(), "standard_answer": q.standard_answer})
    if not unique:
        raise HTTPException(status_code=400, detail="无有效query数据")
    return await eval_label_service.create_label_task(
        db, data.name, data.kb_id, unique, data.top_k, data.description, data.created_by,
    )


@router.get("/label/tasks/{task_id}", response_model=RagEvalLabelTaskResponse)
async def get_label_task(task_id: int, db: AsyncSession = Depends(get_db)):
    task = await eval_label_service.get_label_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="标注任务不存在")
    return task


@router.put("/label/tasks/{task_id}", response_model=RagEvalLabelTaskResponse)
async def update_label_task(task_id: int, data: RagEvalLabelTaskUpdate, db: AsyncSession = Depends(get_db)):
    task = await eval_label_service.update_label_task(db, task_id, data.name, data.description)
    if not task:
        raise HTTPException(status_code=404, detail="标注任务不存在")
    return task


@router.delete("/label/tasks/{task_id}")
async def delete_label_task(task_id: int, db: AsyncSession = Depends(get_db)):
    ok = await eval_label_service.delete_label_task(db, task_id)
    if not ok:
        raise HTTPException(status_code=404, detail="标注任务不存在")
    return {"message": "已删除"}


# ── Label Details ─────────────────────────────────────────

@router.get("/label/tasks/{task_id}/details")
async def list_label_details(
    task_id: int,
    status: str | None = Query(None),
    keyword: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    items, total = await eval_label_service.list_label_details(db, task_id, status, keyword, page, page_size)
    return {"items": [RagEvalLabelDetailResponse.model_validate(d) for d in items], "total": total, "page": page, "page_size": page_size}


@router.get("/label/details/{detail_id}/chunks", response_model=ChunkCandidatesResponse)
async def get_candidate_chunks(detail_id: int, db: AsyncSession = Depends(get_db)):
    result = await eval_label_service.get_candidate_chunks(db, detail_id)
    if result is None:
        raise HTTPException(status_code=404, detail="标注详情不存在")
    return result


@router.put("/label/details/{detail_id}", response_model=RagEvalLabelDetailResponse)
async def save_detail_annotation(detail_id: int, data: RagEvalLabelDetailSave, db: AsyncSession = Depends(get_db)):
    detail = await eval_label_service.save_detail_annotation(db, detail_id, data.standard_chunk_ids, data.annotated_by)
    if not detail:
        raise HTTPException(status_code=404, detail="标注详情不存在")
    return detail


@router.put("/label/tasks/{task_id}/details/batch")
async def batch_save_annotations(task_id: int, data: RagEvalLabelBatchSave, db: AsyncSession = Depends(get_db)):
    saved = await eval_label_service.batch_save_annotations(db, data.annotations, data.annotated_by)
    return {"message": f"已保存 {saved} 条标注", "saved": saved}


# ── Export ────────────────────────────────────────────────

@router.get("/label/tasks/{task_id}/export")
async def export_label_task(task_id: int, format: str = Query("json"), db: AsyncSession = Depends(get_db)):
    task = await eval_label_service.get_label_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="标注任务不存在")
    try:
        result = await eval_label_service.export_label_annotations(db, task_id, format)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if format == "json":
        return {"task": RagEvalLabelTaskResponse.model_validate(task).model_dump(), "items": result}
    elif format == "csv":
        from fastapi.responses import Response
        return Response(content=result, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": f"attachment; filename=label_{task_id}.csv"})
    elif format == "xlsx":
        from fastapi.responses import Response
        return Response(content=result, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=label_{task_id}.xlsx"})
