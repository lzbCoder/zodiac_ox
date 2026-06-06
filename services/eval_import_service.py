import csv
from io import BytesIO, StringIO
from fastapi import UploadFile
from loguru import logger


ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
REQUIRED_COLUMN = "query"
VALID_DIFFICULTIES = {"easy", "medium", "hard"}


def _parse_int_list(raw: str | None) -> list[int] | None:
    """将逗号/分号分隔的字符串解析为整数列表。"""
    if raw is None or not str(raw).strip():
        return None
    result = []
    for part in str(raw).replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            result.append(int(part))
        except ValueError:
            continue
    return result or None


def _validate_row(row: dict, idx: int) -> tuple[bool, str]:
    query = str(row.get(REQUIRED_COLUMN, "")).strip()
    if not query:
        return False, f"第{idx + 1}行缺少「{REQUIRED_COLUMN}」字段"
    diff = str(row.get("difficulty", "medium")).strip().lower()
    if diff and diff not in VALID_DIFFICULTIES:
        return False, f"第{idx + 1}行 difficulty 值无效，仅支持 easy/medium/hard"
    return True, ""


async def parse_upload_file(file: UploadFile) -> list[dict]:
    filename = file.filename or ""
    ext = filename.lower()
    if not any(ext.endswith(e) for e in ALLOWED_EXTENSIONS):
        raise ValueError(f"不支持的文件格式：{ext}，仅支持 Excel(.xlsx/.xls) 或 CSV(.csv)")

    content = await file.read()

    if ext.endswith(".csv"):
        return _parse_csv(content)
    else:
        return _parse_excel(content, ext)


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        ok, err = _validate_row(row, i)
        if not ok:
            raise ValueError(err)
        rows.append({
            "query": str(row["query"]).strip(),
            "standard_answer": str(row.get("standard_answer", "")).strip() or None,
            "standard_doc_ids": _parse_int_list(row.get("standard_doc_ids")),
            "standard_chunk_ids": _parse_int_list(row.get("standard_chunk_ids")),
            "difficulty": str(row.get("difficulty", "medium")).strip().lower() or "medium",
        })
    return rows


def _parse_excel(content: bytes, ext: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl 未安装，无法解析 Excel 文件")

    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise ValueError("Excel 文件为空")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter, [])]

    if REQUIRED_COLUMN not in headers:
        raise ValueError(f"Excel 表头缺少「{REQUIRED_COLUMN}」列，当前表头：{headers}")

    rows = []
    for i, values in enumerate(rows_iter):
        row = dict(zip(headers, values))
        ok, err = _validate_row(row, i + 1)
        if not ok:
            raise ValueError(err)
        rows.append({
            "query": str(row["query"]).strip(),
            "standard_answer": str(row.get("standard_answer", "")).strip() or None,
            "standard_doc_ids": _parse_int_list(row.get("standard_doc_ids")),
            "standard_chunk_ids": _parse_int_list(row.get("standard_chunk_ids")),
            "difficulty": str(row.get("difficulty", "medium")).strip().lower() or "medium",
        })

    wb.close()
    return rows
